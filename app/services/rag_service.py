# app/services/rag_service.py

import os
import shutil
import hashlib
from datetime import datetime
import mimetypes
import re
import traceback 

# Langchain imports
from langchain_community.document_loaders import PyPDFLoader, UnstructuredWordDocumentLoader, UnstructuredODTLoader, TextLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain.docstore.document import Document
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser
from typing import List, Dict, Any, Tuple, Optional

# Local imports
from config import Config
from app.services.llm_service import get_embeddings_llm 

# NEW: Import db and DocumentStatus model
from app import db
from app.models import DocumentStatus

import openpyxl
import pyexcel_ods 

EXCLUDED_EXTENSIONS = {
    '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp', # Images
    '.mp3', '.wav', '.ogg', '.flac', '.aac', # Audio
    '.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv', # Vidéo
    '.zip', '.tar', '.gz', '.rar', '.7z', # Archives
    '.exe', '.dll', '.bin', '.dat', '.so', '.dylib', # Exécutables et bibliothèques
    '.ico', '.db', '.sqlite', '.log', '.bak', '.tmp', # Divers
    '.psd', '.ai', '.eps', # Fichiers Adobe
    '.woff', '.woff2', '.ttf', '.otf', # Fonts
    '.DS_Store', 'thumbs.db', # Fichiers système
    '~$', # Fichiers temporaires Excel/Word
}


class RAGService:
    def __init__(self):
        self.embeddings = get_embeddings_llm() 
        
        self.chroma_path_kb = Config.CHROMA_PATH_KB 
        self.chroma_path_codebase = Config.CHROMA_PATH_CODEBASE 
        self.kb_documents_path = Config.KNOWLEDGE_BASE_DIR 
        self.codebase_path = Config.CODE_BASE_DIR 
        self.processing_cache_path = Config.PROCESSING_CACHE_PATH 

        self._initialize_directories()

        self._any_db_reset = False 

        self.db_kb, kb_was_reset_or_empty = self._get_or_create_vector_store(self.chroma_path_kb)
        self.db_codebase, codebase_was_reset_or_empty = self._get_or_create_vector_store(self.chroma_path_codebase)
        
        if kb_was_reset_or_empty or codebase_was_reset_or_empty:
            self._any_db_reset = True
            print(f"Detected new/empty ChromaDB(s). Clearing processing cache at {self.processing_cache_path} to force full re-ingestion.")
            if os.path.exists(self.processing_cache_path):
                shutil.rmtree(self.processing_cache_path)
            os.makedirs(self.processing_cache_path, exist_ok=True) 


    def _initialize_directories(self):
        os.makedirs(self.kb_documents_path, exist_ok=True)
        os.makedirs(self.codebase_path, exist_ok=True)
        os.makedirs(self.processing_cache_path, exist_ok=True)

    def _get_or_create_vector_store(self, path: str) -> Tuple[Chroma, bool]:
        """Initialise ou charge un vector store ChromaDB.
        Retourne l'instance ChromaDB et un booléen (True si créée/vide, False si chargée avec des données existantes).
        """
        was_reset_or_empty = False
        if not os.path.exists(path) or not os.listdir(path):
            print(f"Creating new ChromaDB at {path}")
            was_reset_or_empty = True
            chroma_db = Chroma(embedding_function=self.embeddings, persist_directory=path)
        else:
            print(f"Loading existing ChromaDB from {path}")
            chroma_db = Chroma(embedding_function=self.embeddings, persist_directory=path)
            if len(chroma_db.get(include=[])['ids']) == 0:
                print(f"Existing ChromaDB at {path} found to be empty. Treating as if newly created.")
                was_reset_or_empty = True 
        return chroma_db, was_reset_or_empty

    def _calculate_file_hash(self, file_path: str) -> str:
        """Calcule le hash MD5 d'un fichier."""
        hasher = hashlib.md5()
        with open(file_path, 'rb') as f:
            while True:
                chunk = f.read(8192)  # Lire par blocs de 8KB
                if not chunk:
                    break
                hasher.update(chunk)
        return hasher.hexdigest()

    def _get_cached_hash(self, file_path: str) -> Optional[str]:
        """Récupère le hash d'un fichier depuis le cache (pour comparaison rapide, non critique)."""
        cache_file_name = hashlib.md5(file_path.encode('utf-8')).hexdigest() + ".hash"
        cache_file = os.path.join(self.processing_cache_path, cache_file_name)
        if os.path.exists(cache_file):
            with open(cache_file, 'r') as f:
                return f.read().strip()
        return None

    def _save_cached_hash(self, file_path: str, file_hash: str):
        """Sauvegarde le hash d'un fichier dans le cache."""
        cache_file_name = hashlib.md5(file_path.encode('utf-8')).hexdigest() + ".hash"
        cache_file = os.path.join(self.processing_cache_path, cache_file_name)
        with open(cache_file, 'w') as f:
                f.write(file_hash)

    def _get_current_files(self, directory: str) -> Dict[str, Dict[str, Any]]:
        """Récupère tous les chemins de fichiers valides dans un répertoire et ses sous-répertoires."""
        current_files_on_disk: Dict[str, Dict[str, Any]] = {}
        for root, _, files in os.walk(directory):
            for file in files:
                file_path = os.path.normpath(os.path.join(root, file)) 
                file_extension = os.path.splitext(file_path)[1].lower()
                
                if file_extension in EXCLUDED_EXTENSIONS or file.startswith('~$'): 
                    print(f"Skipping file due to extension or temp status: {file_path}") 
                    continue
                
                current_files_on_disk[file_path] = {'mtime': os.path.getmtime(file_path), 'size': os.path.getsize(file_path)}
        return current_files_on_disk

    def _load_document(self, file_path: str) -> List[Document]:
        """Charge un document en fonction de son type de fichier."""
        print(f"Loading document: {file_path}")

        file_extension = os.path.splitext(file_path)[1].lower() 
        
        if file_extension in EXCLUDED_EXTENSIONS:
            print(f"Skipping file due to excluded extension (redundant check in loader): {file_path}")
            return []

        try:
            if file_extension == '.pdf':
                loader = PyPDFLoader(file_path)
                docs = loader.load()
            elif file_extension == '.txt':
                loader = TextLoader(file_path, encoding='utf-8')
                docs = loader.load()
            elif file_extension in (".docx", ".doc"):
                loader = UnstructuredWordDocumentLoader(file_path)
                docs = loader.load()
            elif file_extension == ".odt":
                loader = UnstructuredODTLoader(file_path)
                docs = loader.load()
            elif file_extension == ".xlsx":
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                docs_for_file = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_rows_data = []
                    for row in sheet.iter_rows():
                        row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                        sheet_rows_data.append("\t".join(row_values))
                    
                    if sheet_rows_data:
                        sheet_content = f"Feuille: {sheet_name}\n" + "\n".join(sheet_rows_data)
                        metadata = {
                            "source": os.path.abspath(file_path), 
                            "file_type": "kb", 
                            "sheet_name": sheet_name,
                            "is_table_chunk": True 
                        }
                        docs_for_file.append(Document(page_content=sheet_content, metadata=metadata))
                return docs_for_file 
            elif file_extension == ".ods":
                ods_data = pyexcel_ods.get_data(file_path) 
                docs_for_file = [] 
                for sheet_name, table_data in ods_data.items():
                    sheet_rows_data = []
                    for row in table_data:
                        row_values = [str(cell) if cell is not None else "" for cell in row]
                        sheet_rows_data.append("\t".join(row_values)) 
                    
                    if sheet_rows_data:
                        sheet_content = f"Feuille: {sheet_name}\n" + "\n".join(sheet_rows_data)
                        metadata = {
                            "source": os.path.abspath(file_path), 
                            "file_type": "kb", 
                            "sheet_name": sheet_name,
                            "is_table_chunk": True 
                        }
                        docs_for_file.append(Document(page_content=sheet_content, metadata=metadata))
                return docs_for_file 
            else:
                loader = TextLoader(file_path, encoding='utf-8', autodetect_encoding=True)
                docs = loader.load()

            for doc in docs:
                doc.metadata["source"] = os.path.abspath(file_path)
                doc.metadata["file_type"] = "kb" 
            return docs

        except Exception as e:
            print(f"Error loading {file_path}: {e}. Skipping this file.")
            return []

    def _process_documents(self, directory: str, file_type: str, db_instance: Chroma):
        """Generic processor for both KB and Codebase documents."""
        print(f"Processing {file_type.capitalize()} documents from {directory}...")
        current_files_on_disk = self._get_current_files(directory) 
        
        print(f"DEBUG DB: Fetching existing DocumentStatus entries for file_type='{file_type}'.")
        stored_db_status: Dict[str, DocumentStatus] = {
            os.path.normpath(ds.file_path): ds 
            for ds in DocumentStatus.query.filter_by(file_type=file_type).all()
        }
        print(f"DEBUG DB: Found {len(stored_db_status)} existing DocumentStatus entries for '{file_type}'.")


        files_to_add_or_update_paths: set[str] = set()
        files_to_delete_from_db_paths: set[str] = set(stored_db_status.keys()) 

        num_new_files = 0
        num_modified_files = 0
        num_error_files = 0
        num_skipped_files = 0
        num_deleted_files = 0
        
        all_chunks_to_add_in_this_run: List[Document] = [] # Initialized here to ensure it's always bound

        # Phase 1: Identify files to ADD or UPDATE
        for file_path_on_disk in current_files_on_disk.keys(): 
            if file_path_on_disk in files_to_delete_from_db_paths:
                files_to_delete_from_db_paths.remove(file_path_on_disk) 

            current_file_hash = self._calculate_file_hash(file_path_on_disk)
            
            needs_processing = False
            doc_status_entry = stored_db_status.get(file_path_on_disk)

            if self._any_db_reset:
                needs_processing = True
            elif doc_status_entry:
                assert doc_status_entry is not None 
                if current_file_hash != doc_status_entry.file_hash:
                    needs_processing = True
                    num_modified_files += 1
                elif doc_status_entry.status != 'indexed':
                    needs_processing = True
                    if doc_status_entry.status == 'error': num_error_files += 1 
                    elif doc_status_entry.status == 'skipped': num_skipped_files += 1
            else:
                needs_processing = True
                num_new_files += 1
            
            if needs_processing:
                files_to_add_or_update_paths.add(file_path_on_disk)
                self._save_cached_hash(file_path_on_disk, current_file_hash) 


        # Phase 2: Delete documents no longer present on disk
        for file_path_to_delete in files_to_delete_from_db_paths:
            print(f"Deleting removed {file_type.capitalize()} document from ChromaDB: {file_path_to_delete}")
            db_instance.delete(where={"source": file_path_to_delete}) 

            doc_status_entry = stored_db_status.get(file_path_to_delete)
            if doc_status_entry:
                db.session.delete(doc_status_entry)
                num_deleted_files += 1
            
            cache_file_name = hashlib.md5(file_path_to_delete.encode('utf-8')).hexdigest() + ".hash"
            cache_file = os.path.join(self.processing_cache_path, cache_file_name)
            if os.path.exists(cache_file):
                os.remove(cache_file)


        # Phase 3: Load, chunk, and add/update documents in ChromaDB and DocumentStatus
        if files_to_add_or_update_paths:
            for file_path_to_process in files_to_add_or_update_paths:
                status_entry_for_file = stored_db_status.get(file_path_to_process) 
                
                try:
                    current_file_hash = self._calculate_file_hash(file_path_to_process) 
                    loaded_docs: List[Document] = []
                    
                    if file_type == 'kb':
                        loaded_docs = self._load_document(file_path_to_process)
                        final_chunks_for_file = []
                        text_splitter = RecursiveCharacterTextSplitter(chunk_size=Config.CHUNK_SIZE, chunk_overlap=Config.CHUNK_OVERLAP)
                        for doc in loaded_docs:
                            if doc.metadata.get('is_table_chunk'): 
                                final_chunks_for_file.append(doc)
                            else: 
                                split_docs = text_splitter.split_documents([doc])
                                final_chunks_for_file.extend(split_docs)
                        
                        for chunk in final_chunks_for_file:
                            self._add_hierarchical_metadata(chunk, file_path_to_process, file_type)
                        all_chunks_to_add_in_this_run.extend(final_chunks_for_file)


                    elif file_type == 'code':
                        file_extension = os.path.splitext(file_path_to_process)[1].lower()
                        if file_extension in EXCLUDED_EXTENSIONS:
                            print(f"Skipping codebase file {file_path_to_process} as it's detected as binary/unsupported type.")
                            raise ValueError("Binary file detected, cannot read as text.")

                        with open(file_path_to_process, 'r', encoding='utf-8') as f:
                            code_content = f.read()
                        chunks = self._split_code_into_chunks(code_content, file_path_to_process, self._detect_language(file_path_to_process))
                        all_chunks_to_add_in_this_run.extend(chunks)

                    if status_entry_for_file:
                        status_entry_for_file.status = 'indexed'
                        status_entry_for_file.indexed_at = datetime.now()
                        status_entry_for_file.file_hash = current_file_hash
                        status_entry_for_file.last_modified = datetime.fromtimestamp(os.path.getmtime(file_path_to_process))
                        status_entry_for_file.error_message = None
                        db.session.add(status_entry_for_file) 
                    else:
                        new_entry = DocumentStatus(
                            file_path=file_path_to_process,
                            file_type=file_type,
                            status='indexed',
                            last_modified=datetime.fromtimestamp(os.path.getmtime(file_path_to_process)),
                            indexed_at=datetime.now(),
                            file_hash=current_file_hash,
                            error_message=None
                        )
                        db.session.add(new_entry)

                except Exception as e:
                    print(f"Error processing {file_type} file {file_path_to_process}: {e}")
                    traceback.print_exc() 
                    
                    error_status = 'error'
                    error_message = str(e)
                    if "cannot read as text" in str(e).lower() or "codec can't decode" in str(e).lower() or \
                       "file is not a zip file" in str(e).lower(): 
                        error_status = 'skipped'
                        error_message = "File is binary or malformed, skipped for text processing."
                        num_skipped_files += 1 
                    else:
                        num_error_files += 1 

                    current_file_hash_on_error = self._calculate_file_hash(file_path_to_process) 
                    if status_entry_for_file:
                        status_entry_for_file.status = error_status
                        status_entry_for_file.indexed_at = datetime.now()
                        status_entry_for_file.file_hash = current_file_hash_on_error
                        status_entry_for_file.last_modified = datetime.fromtimestamp(os.path.getmtime(file_path_to_process))
                        status_entry_for_file.error_message = error_message
                        db.session.add(status_entry_for_file)
                    else:
                        new_entry = DocumentStatus(
                            file_path=file_path_to_process,
                            file_type=file_type,
                            status=error_status,
                            last_modified=datetime.fromtimestamp(os.path.getmtime(file_path_to_process)),
                            indexed_at=datetime.now(),
                            file_hash=current_file_hash_on_error,
                            error_message=error_message
                        )
                        db.session.add(new_entry)
            
            if all_chunks_to_add_in_this_run:
                sources_to_clear_in_chroma = {os.path.normpath(c.metadata['source']) for c in all_chunks_to_add_in_this_run}
                for source_path in sources_to_clear_in_chroma:
                    db_instance.delete(where={"source": source_path})
                
                db_instance.add_documents(all_chunks_to_add_in_this_run)

        # Summary of processing
        total_files_on_disk = len(current_files_on_disk)
        
        total_indexed_in_db = DocumentStatus.query.filter_by(file_type=file_type, status='indexed').count()
        total_errored_in_db = DocumentStatus.query.filter_by(file_type=file_type, status='error').count()
        total_skipped_in_db = DocumentStatus.query.filter_by(file_type=file_type, status='skipped').count()
        
        print(f"\n--- {file_type.capitalize()} Processing Summary ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ---")
        print(f"Files currently on disk: {total_files_on_disk}")
        print(f"Files known in DocumentStatus (at start of this run): {len(stored_db_status)}")
        
        if num_new_files > 0:
            print(f"  - New files added to DB: {num_new_files}")
        if num_modified_files > 0:
            print(f"  - Modified files re-indexed: {num_modified_files}")
        if len(files_to_delete_from_db_paths) > 0:
            print(f"  - Files deleted from disk: {len(files_to_delete_from_db_paths)} (removed from ChromaDB & DocumentStatus)")
        
        print(f"  - Total chunks added/updated in ChromaDB this run: {len(all_chunks_to_add_in_this_run)}")
        
        print(f"Current DocumentStatus counts (after this run's operations, before final commit):")
        print(f"  - Successfully Indexed: {total_indexed_in_db}")
        print(f"  - With Errors: {total_errored_in_db}")
        print(f"  - Skipped (e.g., binary): {total_skipped_in_db}")
        print(f"--- End {file_type.capitalize()} Summary ---")


    def _process_kb_documents(self):
        self._process_documents(self.kb_documents_path, 'kb', self.db_kb)

    def _process_codebase_documents(self):
        self._process_documents(self.codebase_path, 'code', self.db_codebase)


    def _detect_language(self, file_path: str) -> Optional[str]:
        """Détecte le langage de programmation basé sur l'extension du fichier."""
        extension_map = {
            '.py': 'python',
            '.js': 'javascript',
            '.ts': 'typescript',
            '.jsx': 'javascript', 
            '.tsx': 'typescript', 
            '.java': 'java',
            '.c': 'c',
            '.cpp': 'cpp',
            '.h': 'c_header',
            '.hpp': 'cpp_header',
            '.go': 'go',
            '.rb': 'ruby',
            '.php': 'php',
            '.html': 'html',
            '.css': 'css',
            '.json': 'json', 
            '.xml': 'xml',   
            '.yml': 'yaml', 
            '.yaml': 'yaml',
            '.sh': 'bash', 
            '.md': 'markdown', 
            '.less': 'css', 
            '.svg': 'xml' 
        }
        _, ext = os.path.splitext(file_path)
        return extension_map.get(ext.lower())

    def _split_code_into_chunks(self, code_content: str, file_path: str, language: Optional[str]) -> List[Document]:
        """Divise le code en chunks logiques (fonctions, classes, etc.) et extrait des métadonnées."""
        chunks: List[Document] = []
        lines = code_content.splitlines()
        
        project_name: Optional[str] = None
        try:
            relative_path = os.path.relpath(file_path, self.codebase_path)
            path_parts = os.path.normpath(relative_path).split(os.sep) 
            if len(path_parts) > 0 and not path_parts[0].startswith('.'): 
                project_name = path_parts[0]
        except ValueError:
            pass 

        file_imports_list: List[str] = [] 
        file_imports_str: str = ""

        # --- Langage spécifique: Python ---
        if language == 'python':
            func_class_pattern = re.compile(r"^(async\s+)?(def|class)\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*[:(]")
            import_pattern = re.compile(r"^(?:from\s+([a-zA-Z_][a-zA-Z0-9_\.]*)\s+)?import\s+([a-zA-Z_][a-zA-Z0-9_\.,\s]*)(?:\s+as\s+.*)?")

            for line in lines:
                stripped_line = line.strip()
                if not stripped_line or stripped_line.startswith('#'): 
                    continue
                import_match = import_pattern.match(stripped_line)
                if import_match:
                    if import_match.group(1): 
                        file_imports_list.append(import_match.group(1).split('.')[0]) 
                    else: 
                        file_imports_list.extend([s.strip().split('.')[0] for s in import_match.group(2).split(',')])
                elif not stripped_line.startswith(('from', 'import')): 
                    break
            file_imports_list = list(set(file_imports_list)) 
            file_imports_str = ",".join(file_imports_list)

            current_chunk_content: List[str] = []
            current_chunk_metadata: Dict[str, Any] = {}
            current_entity_start_line = 0

            for i, line in enumerate(lines):
                match = func_class_pattern.match(line)
                if match and current_chunk_content:
                    chunk_imports_to_store = current_chunk_metadata.get("imports_list", []) + file_imports_list
                    chunk_imports_str_for_chunk = ",".join(list(set(chunk_imports_to_store))) 
                    
                    doc_to_add = Document(
                        page_content="\n".join(current_chunk_content),
                        metadata={
                            "language": language,
                            "file_type": "code", 
                            "entity_type": current_chunk_metadata.get("entity_type", "module_top_level"),
                            "entity_name": os.path.basename(file_path),
                            "imports": chunk_imports_str_for_chunk, 
                            "start_line": current_entity_start_line,
                            "end_line": i,
                            "project_name": project_name,
                        }
                    )
                    self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                    chunks.append(doc_to_add)
                    current_chunk_content = []
                    current_chunk_metadata = {}
                
                if match:
                    entity_type = match.group(2) 
                    entity_name = match.group(3)
                    current_chunk_metadata = {
                        "entity_type": "function" if entity_type == "def" else "class",
                        "entity_name": entity_name,
                        "start_line": i,
                        "imports_list": [] 
                    }
                    current_entity_start_line = i
                
                current_chunk_content.append(line)

            if current_chunk_content:
                chunk_imports_to_store = current_chunk_metadata.get("imports_list", []) + file_imports_list
                chunk_imports_str_for_chunk = ",".join(list(set(chunk_imports_to_store)))
                
                doc_to_add = Document(
                    page_content="\n".join(current_chunk_content),
                    metadata={
                        "language": language,
                        "file_type": "code", 
                        "entity_type": current_chunk_metadata.get("entity_type", "module_remaining"),
                        "entity_name": os.path.basename(file_path),
                        "imports": chunk_imports_str_for_chunk, 
                        "start_line": current_entity_start_line,
                        "end_line": len(lines) - 1,
                        "project_name": project_name,
                    }
                )
                self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                chunks.append(doc_to_add)

        # --- Langage spécifique: JavaScript/TypeScript/JSX/TSX ---
        elif language in ['javascript', 'typescript']:
            js_ts_entity_pattern = re.compile(
                r"^(?:export\s+)?(?:async\s+)?(?:function\s+([a-zA-Z_][a-zA-Z0-9_]*)|class\s+([a-zA-Z_][a-zA-Z0-9_]*)|(?:const|let|var)\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*=\s*(?:function|=>)|interface\s+([a-zA-Z_][a-zA-Z0-9_]*)|type\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*=)"
            )
            js_ts_import_pattern = re.compile(r"import\s+(?:\{[^}]+\}|\*\s+as\s+\w+|\w+)\s+from\s+['\"`]([^'\"]+)['\"`]")

            for line in lines:
                stripped_line = line.strip()
                import_match = js_ts_import_pattern.search(stripped_line)
                if import_match:
                    module_path = import_match.group(1)
                    file_imports_list.append(os.path.basename(module_path).split('.')[0]) 
            file_imports_list = list(set(file_imports_list))
            file_imports_str = ",".join(file_imports_list)

            current_chunk_content = []
            current_chunk_metadata = {}
            current_entity_start_line = 0

            for i, line in enumerate(lines):
                match = js_ts_entity_pattern.match(line.strip())
                if match and current_chunk_content: 
                    entity_type = "module_top_level"
                    entity_name = os.path.basename(file_path)
                    
                    if match.group(1): (entity_type, entity_name) = ("function", match.group(1))
                    elif match.group(2): (entity_type, entity_name) = ("class", match.group(2))
                    elif match.group(3): (entity_type, entity_name) = ("variable", match.group(3))
                    elif match.group(4): (entity_type, entity_name) = ("interface", match.group(4))
                    elif match.group(5): (entity_type, entity_name) = ("type", match.group(5))

                    doc_to_add = Document(
                        page_content="\n".join(current_chunk_content),
                        metadata={
                            "language": language,
                            "file_type": "code",
                            "entity_type": current_chunk_metadata.get("entity_type", "module_top_level_segment"),
                            "entity_name": current_chunk_metadata.get("entity_name", os.path.basename(file_path)),
                            "imports": file_imports_str, 
                            "start_line": current_entity_start_line,
                            "end_line": i,
                            "project_name": project_name,
                        }
                    )
                    self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                    chunks.append(doc_to_add)
                    current_chunk_content = []
                    current_chunk_metadata = {
                        "entity_type": entity_type,
                        "entity_name": entity_name,
                        "start_line": i
                    }
                    current_entity_start_line = i
                
                current_chunk_content.append(line)

            if current_chunk_content: 
                doc_to_add = Document(
                    page_content="\n".join(current_chunk_content),
                    metadata={
                        "language": language,
                        "file_type": "code",
                        "entity_type": current_chunk_metadata.get("entity_type", "module_remaining_segment"),
                        "entity_name": os.path.basename(file_path),
                        "imports": file_imports_str,
                        "start_line": current_entity_start_line,
                        "end_line": len(lines) - 1,
                        "project_name": project_name,
                    }
                )
                self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                chunks.append(doc_to_add)

        # --- Langage spécifique: HTML ---
        elif language == 'html':
            html_tag_pattern = re.compile(r"^\s*<(?P<tag_name>[a-zA-Z0-9]+)(?:\s+[^>]*)?>")
            
            html_sources_list: List[str] = []
            script_src_pattern = re.compile(r"<script[^>]*src=['\"]([^'\"]+)['\"][^>]*>")
            link_href_pattern = re.compile(r"<link[^>]*href=['\"]([^'\"]+)['\"][^>]*>")
            
            for line in lines:
                script_match = script_src_pattern.search(line)
                if script_match:
                    html_sources_list.append(script_match.group(1))
                link_match = link_href_pattern.search(line)
                if link_match:
                    html_sources_list.append(link_match.group(1))
            file_imports_str = ",".join(list(set(html_sources_list)))

            current_chunk_content = []
            current_tag = "document_root" 
            current_entity_start_line = 0

            for i, line in enumerate(lines):
                match = html_tag_pattern.match(line)
                if match:
                    tag_name = match.group("tag_name")
                    if tag_name in ['html', 'head', 'body', 'div', 'section', 'article', 'nav', 'header', 'footer', 'main'] and current_chunk_content:
                        doc_to_add = Document(
                            page_content="\n".join(current_chunk_content),
                            metadata={
                                "language": language,
                                "file_type": "code",
                                "entity_type": "html_tag_block",
                                "entity_name": current_tag,
                                "imports": file_imports_str,
                                "start_line": current_entity_start_line,
                                "end_line": i,
                                "project_name": project_name,
                            }
                        )
                        self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                        chunks.append(doc_to_add)
                        current_chunk_content = []
                        current_entity_start_line = i
                    current_tag = tag_name 
                
                current_chunk_content.append(line)
            
            if current_chunk_content: 
                doc_to_add = Document(
                    page_content="\n".join(current_chunk_content),
                    metadata={
                        "language": language,
                        "file_type": "code",
                        "entity_type": "html_tag_remaining",
                        "entity_name": current_tag,
                        "imports": file_imports_str,
                        "start_line": current_entity_start_line,
                        "end_line": len(lines) - 1,
                        "project_name": project_name,
                    }
                )
                self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                chunks.append(doc_to_add)

        # --- Langage spécifique: CSS / LESS ---
        elif language == 'css':
            css_rule_pattern = re.compile(r"^\s*([a-zA-Z0-9\s\.\#\:\,\-\>\[\]\(\)\"\'=\~]+)\s*\{") 
            
            css_imports_list: List[str] = []
            at_import_pattern = re.compile(r"@import\s+['\"]([^'\"]+)['\"];")
            for line in lines:
                import_match = at_import_pattern.search(line)
                if import_match:
                    css_imports_list.append(import_match.group(1))
            file_imports_str = ",".join(list(set(css_imports_list)))

            current_chunk_content = []
            current_selector = "document_styles"
            current_rule_start_line = 0

            for i, line in enumerate(lines):
                match = css_rule_pattern.match(line)
                if match:
                    selector = match.group(1).strip()
                    if current_chunk_content:
                        doc_to_add = Document(
                            page_content="\n".join(current_chunk_content),
                            metadata={
                                "language": language,
                                "file_type": "code",
                                "entity_type": "css_rule_block",
                                "entity_name": current_selector,
                                "imports": file_imports_str,
                                "start_line": current_rule_start_line,
                                "end_line": i,
                                "project_name": project_name,
                            }
                        )
                        self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                        chunks.append(doc_to_add)
                        current_chunk_content = []
                        current_rule_start_line = i
                    current_selector = selector
                
                current_chunk_content.append(line)
            
            if current_chunk_content: 
                doc_to_add = Document(
                    page_content="\n".join(current_chunk_content),
                    metadata={
                        "language": language,
                        "file_type": "code",
                        "entity_type": "css_rule_remaining",
                        "entity_name": current_selector,
                        "imports": file_imports_str,
                        "start_line": current_rule_start_line,
                        "end_line": len(lines) - 1,
                        "project_name": project_name,
                    }
                )
                self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                chunks.append(doc_to_add)
        
        # --- Langage spécifique: YAML ---
        elif language == 'yaml':
            yaml_top_level_key_pattern = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_-]*:\s*.*")
            
            current_chunk_content = []
            current_key = "document_root" 
            current_block_start_line = 0

            for i, line in enumerate(lines):
                match = yaml_top_level_key_pattern.match(line)
                if match and not line.startswith(" ") and current_chunk_content: 
                    key_name = line.split(':')[0].strip()
                    doc_to_add = Document(
                        page_content="\n".join(current_chunk_content),
                        metadata={
                            "language": language,
                            "file_type": "code",
                            "entity_type": "yaml_block",
                            "entity_name": current_key,
                            "imports": "", 
                            "start_line": current_block_start_line,
                            "end_line": i,
                            "project_name": project_name,
                        }
                    )
                    self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                    chunks.append(doc_to_add)
                    current_chunk_content = []
                    current_block_start_line = i
                    current_key = key_name
                
                current_chunk_content.append(line)
            
            if current_chunk_content: 
                doc_to_add = Document(
                    page_content="\n".join(current_chunk_content),
                    metadata={
                        "language": language,
                        "file_type": "code",
                        "entity_type": "yaml_block_remaining", 
                        "entity_name": current_key,
                        "imports": "",
                        "start_line": current_block_start_line,
                        "end_line": len(lines) - 1,
                        "project_name": project_name,
                    }
                )
                self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                chunks.append(doc_to_add)

        # --- Langage spécifique: Markdown ---
        elif language == 'markdown':
            markdown_heading_pattern = re.compile(r"^(#+)\s*(.*)")
            
            current_chunk_content = []
            current_heading = "document_top" 
            current_section_start_line = 0

            for i, line in enumerate(lines):
                match = markdown_heading_pattern.match(line)
                if match and current_chunk_content:
                    heading_level = len(match.group(1))
                    heading_text = match.group(2).strip()
                    doc_to_add = Document(
                        page_content="\n".join(current_chunk_content),
                        metadata={
                            "language": language,
                            "file_type": "code", 
                            "entity_type": f"markdown_heading_level_{heading_level}",
                            "entity_name": current_heading, 
                            "imports": "", 
                            "start_line": current_section_start_line,
                            "end_line": i,
                            "project_name": project_name,
                        }
                    )
                    self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                    chunks.append(doc_to_add)
                    current_chunk_content = []
                    current_section_start_line = i
                    current_heading = heading_text 
                
                current_chunk_content.append(line)
            
            if current_chunk_content: 
                doc_to_add = Document(
                    page_content="\n".join(current_chunk_content),
                    metadata={
                        "language": language,
                        "file_type": "code",
                        "entity_type": f"markdown_section_remaining", 
                        "entity_name": current_heading,
                        "imports": "",
                        "start_line": current_section_start_line,
                        "end_line": len(lines) - 1,
                        "project_name": project_name,
                    }
                )
                self._add_hierarchical_metadata(doc_to_add, file_path, 'code')
                chunks.append(doc_to_add)

        # --- Fallback pour les langues non implémentées spécifiquement ---
        else:
            print(f"Warning: Advanced code splitting not implemented for {language}. Falling back to general text chunks.")
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=Config.CHUNK_SIZE, 
                chunk_overlap=Config.CHUNK_OVERLAP,
                length_function=len,
                add_start_index=True,
            )
            base_doc = Document(page_content=code_content, metadata={
                "language": language,
                "file_type": "code", 
                "entity_type": "module_generic_chunk", 
                "entity_name": os.path.basename(file_path),
                "imports": "", 
                "start_line": 0,
                "end_line": len(lines) - 1,
                "project_name": project_name,
            })
            chunks_from_splitter = text_splitter.split_documents([base_doc])
            for chunk in chunks_from_splitter:
                chunk.metadata['source'] = os.path.abspath(file_path)
                self._add_hierarchical_metadata(chunk, file_path, 'code')
                chunks.append(chunk)

        return chunks

    def update_vector_store(self):
        """Met à jour le vector store en traitant les nouveaux/modifiés/supprimés documents.
        Cette méthode doit être appelée dans un app_context Flask pour les opérations DB.
        """
        print("Starting RAG service update...")
        try:
            self._process_kb_documents()
            self._process_codebase_documents()
            db.session.commit() # Commit all changes at the end of the update
            print("DEBUG DB: All DocumentStatus changes committed.")
        except Exception as e:
            print(f"FATAL ERROR during RAG service update: {e}")
            traceback.print_exc()
            db.session.rollback() # Rollback if a fatal error occurred
            print("DEBUG DB: DocumentStatus changes rolled back due to error.")
        print("RAG service update complete.")

    def _add_hierarchical_metadata(self, doc: Document, file_path: str, file_type: str): 
        """Ajoute les métadonnées de dossier et de nom de fichier/titre au chunk."""
        base_dir = self.kb_documents_path if file_type == 'kb' else self.codebase_path
        
        absolute_file_path = os.path.abspath(file_path)
        doc.metadata['source'] = os.path.normpath(absolute_file_path) 

        doc.metadata['document_path_relative'] = os.path.normpath(os.path.relpath(absolute_file_path, base_dir))

        path_components = os.path.normpath(os.path.relpath(absolute_file_path, base_dir)).split(os.sep)

        MAX_FOLDER_LEVELS = 3 

        for i, component in enumerate(path_components[:-1]): 
            if i < MAX_FOLDER_LEVELS:
                doc.metadata[f'folder_level_{i+1}'] = component
            if i == len(path_components) - 2: 
                doc.metadata['last_folder_name'] = component

        doc.metadata['file_name'] = os.path.basename(file_path) 
        doc.metadata['file_type'] = file_type 
        
        if file_type == 'code' and len(path_components) > 0 and path_components[0] and not path_components[0].startswith('.'): 
            doc.metadata['project_name'] = path_components[0]
        else: 
            doc.metadata['project_name'] = None


        if 'title' not in doc.metadata or not doc.metadata['title']: 
            doc.metadata['document_title'] = os.path.splitext(os.path.basename(file_path))[0]
        else:
            doc.metadata['document_title'] = doc.metadata['title'] 

    def get_kb_db_instance(self) -> Chroma: # NEW: Return Chroma instance directly
        return self.db_kb

    def get_codebase_db_instance(self) -> Chroma: # NEW: Return Chroma instance directly
        return self.db_codebase


# This block is for direct testing of RAGService outside Flask app.
# It requires a minimal Flask app context setup for SQLAlchemy.
if __name__ == "__main__":
    from flask import Flask
    os.environ['DB_USER'] = os.environ.get('DB_USER', 'dev_user')
    os.environ['DB_PASSWORD'] = os.environ.get('DB_PASSWORD', 'dev_password')
    os.environ['DB_HOST'] = os.environ.get('DB_HOST', 'localhost')
    os.environ['DB_PORT'] = os.environ.get('DB_PORT', '5432')
    os.environ['DB_NAME'] = os.environ.get('DB_NAME', 'mon_premier_rag_db')
    os.environ['LMSTUDIO_API_KEY'] = os.environ.get('LMSTUDIO_API_KEY', 'lm-studio') 
    os.environ['LMSTUDIO_CHAT_MODEL'] = os.environ.get('LMSTUDIO_CHAT_MODEL', 'Llama-3.1-8B-UltraLong-4M-Instruct-Q4_K_M')

    temp_app = Flask(__name__)
    temp_app.config.from_object(Config)
    db.init_app(temp_app)

    with temp_app.app_context():
        db.create_all()

    print("Initializing RAGService for direct testing...")
    rag_service = RAGService()
    print("Updating vector stores...")
    
    with temp_app.app_context():
        try:
            rag_service.update_vector_store()
            db.session.commit() 
        except Exception as e:
            print(f"Error during RAG service update in test run: {e}")
            db.session.rollback() 
            traceback.print_exc()


    with temp_app.app_context():
        print("\nKB DocumentStatus in DB:")
        kb_statuses = DocumentStatus.query.filter_by(file_type='kb').all()
        if kb_statuses:
            for status_entry in kb_statuses:
                print(f"- {status_entry.file_path} | Status: {status_entry.status} | Last Ingested: {status_entry.indexed_at} | Error: {status_entry.error_message}")
        else:
            print("No KB DocumentStatus entries found.")

        print("\nCodebase DocumentStatus in DB:")
        code_statuses = DocumentStatus.query.filter_by(file_type='code').all()
        if code_statuses:
            for status_entry in code_statuses:
                print(f"- {status_entry.file_path} | Status: {status_entry.status} | Last Ingested: {status_entry.indexed_at} | Error: {status_entry.error_message}")
        else:
            print("No Codebase DocumentStatus entries found.")